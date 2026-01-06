# type: ignore
from fastapi import FastAPI, UploadFile, File, HTTPException,Form
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pandas as pd
import numpy as np
import tempfile
import os
import re
from typing import Dict, Iterable, List, Optional, Tuple
from starlette.background import BackgroundTask
from validation import normalize_dataframe
from error_summary import write_error_summary_sheet, write_highlight_summary_sheet

app = FastAPI(title="Mapping Normalization API")

# Allow local dev from the Next.js frontend.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)



def normalize_report_name(name: str) -> str:
    """Normalize report/file names for matching."""
    return re.sub(r"\s+", " ", os.path.splitext(str(name).strip().lower())[0])


def normalize_column_name(name: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(name or "").strip().lower())
    cleaned = re.sub(r"[^a-z0-9 #/_-]", "", cleaned)
    return cleaned


def parse_mapping(upload: UploadFile) -> List[dict]:
    """
    Read mapping using the strict format:
    output_col, report_name, column_name, possible_variations, default_value
    """
    try:
        mapping_df = pd.read_excel(upload.file, engine="openpyxl")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to read mapping file: {exc}")

    normalized_cols = {normalize_column_name(c): c for c in mapping_df.columns}
    required = {"output_col", "report_name", "column_name", "possible_variations", "default_value"}
    if not required.issubset(set(normalized_cols.keys())):
        missing = required - set(normalized_cols.keys())
        raise HTTPException(
            status_code=400,
            detail=f"Mapping file missing required columns: {', '.join(sorted(missing))}",
        )
    default_col_name = normalized_cols["default_value"]

    rules: List[dict] = []
    for _, row in mapping_df.iterrows():
        output_col = str(row.get(normalized_cols["output_col"], "")).strip()
        report_name = str(row.get(normalized_cols["report_name"], "")).strip()
        column_name = str(row.get(normalized_cols["column_name"], "")).strip()
        variations_raw = row.get(normalized_cols["possible_variations"], "")
        variations_raw = "" if pd.isna(variations_raw) else str(variations_raw).strip()
        default_raw = row.get(default_col_name, "")
        default_value = "" if pd.isna(default_raw) else str(default_raw).strip()
        # Default-only rows are allowed only when output_col is present AND report_name/column_name are both empty.
        if not output_col:
            continue
        # Allow default-driven rows when output_col exists and either report/column are missing
        # but default_value is provided. Otherwise skip.
        if (not report_name or not column_name) and default_value == "":
            continue
        variations = (
            [v.strip() for v in variations_raw.replace(";", ",").split(",") if v.strip()]
            if variations_raw
            else []
        )
        rules.append(
            {
                "output_col": output_col,
                "report_name": report_name,
                "report_key": normalize_report_name(report_name) if report_name else "",
                "column_name": column_name,
                "variations": variations,
                "default_value": default_value,
            }
        )

    if not rules:
        raise HTTPException(status_code=400, detail="Mapping file contained no usable rows.")
    return rules

def parse_template(upload: UploadFile) -> List[str]:
    """
    Read template.xlsx to extract the desired output column order.
    """
    try:
        template_df = pd.read_excel(upload.file, engine="openpyxl")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to read template file: {exc}")

    if "output_col" not in template_df.columns:
        raise HTTPException(status_code=400, detail="Template file must contain 'output_col' column")

    template_order = [str(val).strip() for val in template_df["output_col"] if pd.notna(val)]
    if not template_order:
        raise HTTPException(status_code=400, detail="Template file contains no output_col values")
    return template_order


def make_unique(columns: List[str]) -> List[str]:
    """Ensure column names are unique by appending counters."""
    seen = {}
    unique_cols = []
    for col in columns:
        base = col
        if base not in seen:
            seen[base] = 0
            unique_cols.append(base)
        else:
            seen[base] += 1
            unique_cols.append(f"{base}_{seen[base]}")
    return unique_cols


def find_header_row(raw_df: pd.DataFrame, candidate_names: set) -> int:
    """
    Auto-detect header row inspired by context/excel_merger.py:
      - Scores rows based on non-null, string, uniqueness, presence of candidate mapping names,
        and common column-like words.
    """
    max_rows = min(25, len(raw_df))
    best_row = 0
    best_score = -1
    common_column_words = [
        "name",
        "date",
        "type",
        "status",
        "id",
        "number",
        "rent",
        "unit",
        "tenant",
        "space",
        "tenant name",
        "balance",
        "paid",
        "total",
        "fee",
        "charge",
        "billing",
        "city",
        "age",
    ]

    for idx in range(max_rows):
        row = raw_df.iloc[idx]
        if row.isna().all():
            continue

        non_null_count = row.notna().sum()
        string_count = sum(isinstance(val, str) and len(str(val).strip()) > 0 for val in row)
        unique_count = len(row.dropna().unique())
        numeric_count = sum(isinstance(val, (int, float)) and not pd.isna(val) for val in row)

        # Column-like detection
        column_like_count = 0
        for val in row:
            if isinstance(val, str):
                val_lower = val.lower().strip()
                if len(val_lower) < 50 and any(word in val_lower for word in common_column_words):
                    column_like_count += 1
                elif len(val_lower.split()) <= 4 and len(val_lower) < 50:
                    column_like_count += 0.5

        # Candidate mapping name bonus
        row_vals = {str(v).strip().lower() for v in row if pd.notna(v) and str(v).strip()}
        candidate_overlap = len(row_vals & candidate_names)

        if non_null_count < 1:
            continue

        score = (string_count * 3) + (unique_count * 2) + non_null_count
        score += column_like_count * 5
        score += candidate_overlap * 15  # strong signal if mapping names appear

        if string_count >= non_null_count * 0.8:
            score += 10
        if unique_count >= non_null_count * 0.9:
            score += 15
        if numeric_count > string_count:
            score -= 10
        if non_null_count <= 2:
            score -= 20

        if score > best_score:
            best_score = score
            best_row = idx

    return best_row


def clean_dataframe(raw_df: pd.DataFrame, mapping_rules: List[dict]) -> pd.DataFrame:
    """
    Clean messy inputs where headers or metadata occupy top rows.
    Strategy:
      - Read with header=None.
      - Detect the header row using scoring informed by mapping names and column-like heuristics.
      - Use that row as header and drop rows above it.
      - Drop rows/columns that are entirely empty after header assignment.
    """
    # Build set of candidate header tokens from mapping columns and variations.
    candidate_names = set()
    for rule in mapping_rules:
        candidate_names.add(rule["column_name"].lower())
        candidate_names.add(rule["output_col"].lower())
        for var in rule.get("variations", []):
            candidate_names.add(var.lower())

    header_index = find_header_row(raw_df, candidate_names)

    # Set header and data
    header_row = raw_df.iloc[header_index].fillna("").astype(str).str.strip().tolist()
    header_row = [h if h else "unnamed" for h in header_row]
    header_row = make_unique(header_row)

    data_df = raw_df.iloc[header_index + 1 :].copy()
    data_df.columns = header_row[: len(data_df.columns)]
    data_df = data_df.reset_index(drop=True)

    # Drop rows that are entirely empty/blank after header assignment.
    def row_is_empty(row):
        return all((pd.isna(v) or str(v).strip() == "") for v in row)

    data_df = data_df[~data_df.apply(row_is_empty, axis=1)]
    data_df = data_df.reset_index(drop=True)

    # Drop columns that are entirely empty/NaN/blank
    def is_empty_series(s: pd.Series) -> bool:
        return s.isna().all() or all(str(v).strip() == "" for v in s)

    data_df = data_df[[col for col in data_df.columns if not is_empty_series(data_df[col])]]
    return data_df


def read_input_file(upload: UploadFile, mapping_rules: List[dict]) -> pd.DataFrame:
    """
    Read an uploaded CSV or Excel file into a cleaned DataFrame.
    Handles messy headers by scanning for the likely header row.
    """
    filename = upload.filename.lower()
    try:
        if filename.endswith(".csv"):
            raw_df = pd.read_csv(upload.file, header=None)
        elif filename.endswith(".xls"):
            raw_df = pd.read_excel(upload.file, engine="xlrd", header=None)
        else:
            raw_df = pd.read_excel(upload.file, engine="openpyxl", header=None)
        return clean_dataframe(raw_df, mapping_rules)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to read input file {upload.filename}: {exc}")



def find_matching_column(df: pd.DataFrame, base_name: str, variations: Iterable[str]) -> Optional[str]:
    """
    Resolve the best matching column in df for the given base_name/variations.
    Strategy:
    - Exact raw match.
    - Exact normalized match.
    - Variation raw/normalized match.
    - Partial/fuzzy contains on normalized names.
    Returns the original column name or None.
    """
    if df is None or df.empty:
        return None

    base_name = base_name or ""
    variation_list = [v for v in variations if v] if variations else []
    normalized_lookup = {normalize_column_name(col): col for col in df.columns}

    if base_name in df.columns:
        return base_name

    base_norm = normalize_column_name(base_name)
    if base_norm in normalized_lookup:
        return normalized_lookup[base_norm]

    for v in variation_list:
        if v in df.columns:
            return v
        v_norm = normalize_column_name(v)
        if v_norm in normalized_lookup:
            return normalized_lookup[v_norm]

    candidates = []
    for norm, original in normalized_lookup.items():
        score = 0
        if base_norm and base_norm in norm:
            score += 2
        for v in variation_list:
            v_norm = normalize_column_name(v)
            if v_norm and v_norm in norm:
                score += 1
        if score > 0:
            candidates.append((score, original))

    if candidates:
        candidates.sort(key=lambda x: (-x[0], x[1]))
        return candidates[0][1]

    return None


def select_join_key(
    dataframes: Dict[str, pd.DataFrame], normalized_columns: Dict[str, Dict[str, str]]
) -> Tuple[Optional[str], Dict[str, str]]:
    """
    Detect a join key across source dataframes.
    Returns (normalized_key, per_report_original_name) or (None, {}).
    """
    if not dataframes:
        return None, {}

    presence: Dict[str, List[str]] = {}
    for report, norm_map in normalized_columns.items():
        for norm in norm_map.keys():
            presence.setdefault(norm, []).append(report)

    candidates = {norm: reps for norm, reps in presence.items() if len(reps) >= 2}
    if not candidates:
        return None, {}

    def score_norm(norm: str, reports: List[str]) -> float:
        name_bonus = 0
        if any(k in norm for k in ["id", "email", "address", "space", "unit"]):
            name_bonus += 1
        uniques = []
        coverages = []
        for rep in reports:
            df = dataframes[rep]
            col = normalized_columns[rep][norm]
            series = df[col]
            non_null = series.dropna()
            if len(series) == 0:
                continue
            uniques.append(non_null.nunique() / max(len(non_null), 1))
            coverages.append(len(non_null) / max(len(series), 1))
        avg_unique = np.mean(uniques) if uniques else 0
        avg_cov = np.mean(coverages) if coverages else 0
        return (len(reports) * 2) + avg_unique + avg_cov + name_bonus

    best_norm = None
    best_score = -1
    for norm, reps in candidates.items():
        sc = score_norm(norm, reps)
        if sc > best_score:
            best_score = sc
            best_norm = norm

    if not best_norm:
        return None, {}

    per_report = {
        rep: normalized_columns[rep][best_norm]
        for rep in presence.get(best_norm, [])
        if best_norm in normalized_columns[rep]
    }
    return best_norm, per_report


def collect_master_keys(
    dataframes: Dict[str, pd.DataFrame], join_key_norm: Optional[str], per_report_key: Dict[str, str]
) -> List:
    if not join_key_norm:
        return []
    keys = set()
    for report, df in dataframes.items():
        col = per_report_key.get(report)
        if not col or col not in df.columns:
            continue
        valid_keys = df[col].dropna()
        valid_keys = [v for v in valid_keys if str(v).strip() != ""]
        keys.update(valid_keys)
    return list(keys)


def is_blank_value(value) -> bool:
    """Treat NaN/None/empty-string as blank for default handling."""
    try:
        return pd.isna(value) or str(value).strip() == ""
    except Exception:
        return False


def is_empty_mapping_target(report_key: str, base_col: str) -> Tuple[bool, bool, bool]:
    """
    Determine whether mapping targets are provided.
    Returns (default_only, has_report, has_column).
    """
    has_report = bool(str(report_key).strip())
    has_column = bool(str(base_col).strip())
    default_only = not has_report and not has_column
    return default_only, has_report, has_column


def build_output_from_mapping(
    mapping_rules: List[dict],
    dataframes: Dict[str, pd.DataFrame],
    normalized_columns: Dict[str, Dict[str, str]],
    template_order: Optional[List[str]],
    owner_name: str
) -> pd.DataFrame:
    """Construct the final output using mapping rules and row-level matching."""
    if not mapping_rules:
        raise ValueError("Mapping rules not loaded.")
    if not dataframes:
        raise ValueError("No data files provided.")

    join_key_norm, per_report_key = select_join_key(dataframes, normalized_columns)
    master_keys = collect_master_keys(dataframes, join_key_norm, per_report_key)
    if master_keys:
        master_keys = sorted(master_keys, key=lambda x: str(x))

    resolved_columns: Dict[str, Dict[Tuple[str, str], Optional[str]]] = {}

    def resolve_column(report: str, base: str, variations: List[str]) -> Optional[str]:
        cache = resolved_columns.setdefault(report, {})
        key = (base, ",".join(variations))
        if key in cache:
            return cache[key]
        df = dataframes.get(report)
        col = find_matching_column(df, base, variations) if df is not None else None
        cache[key] = col
        return col

    output_rows = []
    owner_col_norms = {
        normalize_column_name("Owner"),
        normalize_column_name("name"),
    }

    if master_keys:
        for entity_key in master_keys:
            row_out: Dict[str, object] = {}
            for rule in mapping_rules:
                out_col = rule["output_col"]
                report_key = rule["report_key"]
                base_col = rule["column_name"]
                variations = rule.get("variations", [])
                default_value = rule.get("default_value", "")
                if normalize_column_name(out_col) in owner_col_norms:
                    row_out[out_col] = owner_name
                    continue
                default_only, has_report, has_column = is_empty_mapping_target(report_key, base_col)
                # Critical rule: when both report_name and column_name are empty, always use default_value.
                if default_only:
                    row_out[out_col] = default_value
                    continue
                # Only attempt file-based extraction when both targets are present; otherwise use default.
                if not (has_report and has_column):
                    row_out[out_col] = default_value
                    continue
                df = dataframes.get(report_key)
                value = ""
                if df is not None and base_col and report_key:
                    join_col = per_report_key.get(report_key)
                    target_col = resolve_column(report_key, base_col, variations)
                    if join_col and target_col and join_col in df.columns and target_col in df.columns:
                        matches = df[df[join_col] == entity_key]
                        if not matches.empty:
                            value = matches[target_col].iloc[0]
                    elif target_col and target_col in df.columns and not df.empty:
                        value = df[target_col].iloc[0]
                if is_blank_value(value) and default_value != "" and (not base_col or not report_key):
                    value = default_value
                row_out[out_col] = value
            output_rows.append(row_out)
    else:
        max_len = max(len(df) for df in dataframes.values())
        for idx in range(max_len):
            row_out = {}
            for rule in mapping_rules:
                out_col = rule["output_col"]
                report_key = rule["report_key"]
                base_col = rule["column_name"]
                variations = rule.get("variations", [])
                default_value = rule.get("default_value", "")
                if normalize_column_name(out_col) in owner_col_norms:
                    row_out[out_col] = owner_name
                    continue
                default_only, has_report, has_column = is_empty_mapping_target(report_key, base_col)
                # Critical rule: when both report_name and column_name are empty, always use default_value.
                if default_only:
                    row_out[out_col] = default_value
                    continue
                # Only attempt file-based extraction when both targets are present; otherwise use default.
                if not (has_report and has_column):
                    row_out[out_col] = default_value
                    continue
                df = dataframes.get(report_key)
                value = ""
                if df is not None and len(df) > idx and base_col and report_key:
                    target_col = resolve_column(report_key, base_col, variations)
                    if target_col and target_col in df.columns:
                        value = df[target_col].iloc[idx]
                if is_blank_value(value) and default_value != "" and (not base_col or not report_key):
                    value = default_value
                row_out[out_col] = value
            output_rows.append(row_out)

    output_df = pd.DataFrame(output_rows)
    if not any(normalize_column_name(c) == normalize_column_name("Owner") for c in output_df.columns):
        output_df["Owner"] = owner_name
    if not any(normalize_column_name(c) == normalize_column_name("name") for c in output_df.columns):
        output_df["name"] = owner_name

    if template_order:
        for col in template_order:
            if col not in output_df.columns:
                output_df[col] = ""
        extras = [c for c in output_df.columns if c not in template_order]
        output_df = output_df[template_order + extras]
    else:
        mapping_order = [rule["output_col"] for rule in mapping_rules]
        extras = [c for c in output_df.columns if c not in mapping_order]
        output_df = output_df[mapping_order + extras]

    output_df = coerce_column_types(output_df)
    return output_df


def coerce_column_types(df: pd.DataFrame) -> pd.DataFrame:
    """
    Copy values through without coercing types. We intentionally avoid
    converting to booleans/numbers/dates to preserve the source values.
    """
    return df.copy()


@app.post("/process")
async def process_files(
    mapping: UploadFile = File(...),
    files: List[UploadFile] = File(...),
    template: Optional[UploadFile] = File(None),
    owner_name: str = Form(...),
    migration_date: str = Form(...),
):
    if not owner_name or owner_name.strip() == "":
        raise HTTPException(status_code=400, detail="Owner name is required.")
    if not migration_date or migration_date.strip() == "":
        raise HTTPException(status_code=400, detail="Migration date is required.")
    mig_date = pd.to_datetime(migration_date,errors="coerce")
    if pd.isna(mig_date):
        raise HTTPException(status_code=400, detail="Migration date is invalid.")
    if not files:
        raise HTTPException(status_code=400, detail="At least one data file is required.")

    # Persist mapping upload to disk for reuse in parsing and validation.
    mapping_suffix = os.path.splitext(mapping.filename or "")[1] or ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=mapping_suffix) as map_tmp:
        mapping.file.seek(0)
        map_tmp.write(mapping.file.read())
        mapping_tmp_path = map_tmp.name
    mapping.file.seek(0)

    mapping_rules = parse_mapping(mapping)
    template_order = parse_template(template) if template else None

    dataframes: Dict[str, pd.DataFrame] = {}
    normalized_columns: Dict[str, Dict[str, str]] = {}
    input_filenames: List[str] = []
    for upload in files:
        input_filenames.append(upload.filename or "unknown")
        cleaned_df = read_input_file(upload, mapping_rules)
        report_key = normalize_report_name(upload.filename)
        dataframes[report_key] = cleaned_df
        normalized_columns[report_key] = {normalize_column_name(c): c for c in cleaned_df.columns}

    merged_df = build_output_from_mapping(
        mapping_rules,
        dataframes,
        normalized_columns,
        template_order,
        owner_name.strip(),
    )

    # Validate/normalize merged output.
    validated_df, invalid_cells, highlight_cells, invalid_reasons = normalize_dataframe(
        merged_df, mapping_tmp_path, mig_date=migration_date
    )

    # Sort after validation while preserving correct highlight indices.
    if "First Name" in validated_df.columns:
        validated_df = validated_df.copy()
        validated_df["_orig_index"] = validated_df.index
        validated_df["_first_name_norm"] = (
            validated_df["First Name"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
        )
        validated_df["_first_name_blank"] = validated_df["_first_name_norm"].isin(
            ["", "nan", "none", "null", "na", "n/a"]
        )
        validated_df = validated_df.sort_values(
            by=["_first_name_blank", "_first_name_norm"],
            na_position="last",
        ).reset_index(drop=True)
        index_map = {
            int(row["_orig_index"]): int(idx) for idx, row in validated_df.iterrows()
        }
        validated_df = validated_df.drop(columns=["_orig_index", "_first_name_norm", "_first_name_blank"])

        remapped_invalid = {}
        for col, idx_list in invalid_cells.items():
            remapped = [index_map[i] for i in idx_list if i in index_map]
            if remapped:
                remapped_invalid[col] = remapped
        invalid_cells = remapped_invalid

        remapped_highlight = {"red": {}, "blue": {}, "dark_red": {}, "yellow": {}}
        for color, col_map in highlight_cells.items():
            remapped_cols = {}
            for col, idx_list in col_map.items():
                remapped = [index_map[i] for i in idx_list if i in index_map]
                if remapped:
                    remapped_cols[col] = remapped
            if remapped_cols:
                remapped_highlight[color] = remapped_cols
        highlight_cells = remapped_highlight

        remapped_reasons = []
        for reason in invalid_reasons:
            old_idx = reason.get("row_index")
            if old_idx in index_map:
                updated = dict(reason)
                updated["row_index"] = index_map[old_idx]
                remapped_reasons.append(updated)
        invalid_reasons = remapped_reasons

    # Write to a temporary Excel file and return it.
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
        validated_df.to_excel(tmp_path, index=False, engine="openpyxl")

    # Highlight invalid cells in red and informational highlights in their own colors.
    if invalid_cells or any(col_map for col_map in highlight_cells.values()):
        wb = load_workbook(tmp_path)
        ws = wb.active
        ws.title = "Final Sheet (View Only)"
        header_to_col = {cell.value: cell.column for cell in ws[1]}
        red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        blue_fill = PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
        dark_red_fill = PatternFill(start_color="FF9C0006", end_color="FF9C0006", fill_type="solid")

        for col, idx_list in invalid_cells.items():
            col_num = header_to_col.get(col)
            if not col_num:
                continue
            for idx in idx_list:
                excel_row = idx + 2  # 1-based Excel rows, accounting for header row
                ws.cell(row=excel_row, column=col_num).fill = red_fill

        highlight_order = ["red", "blue", "yellow", "dark_red"]
        for color in highlight_order:
            col_map = highlight_cells.get(color)
            if not col_map:
                continue
            fill = red_fill if color == "red" else blue_fill
            if color == "yellow":
                fill = yellow_fill
            if color == "dark_red":
                fill = dark_red_fill
            for col, idx_list in col_map.items():
                col_num = header_to_col.get(col)
                if not col_num:
                    continue
                for idx in idx_list:
                    excel_row = idx + 2
                    ws.cell(row=excel_row, column=col_num).fill = fill

        # Write error and calculation summary sheets.
        if invalid_cells:
            invalid_lookup = {col: set(rows) for col, rows in invalid_cells.items()}
            error_reasons = [
                reason
                for reason in invalid_reasons
                if reason.get("column") in invalid_lookup
                and reason.get("row_index") in invalid_lookup[reason.get("column")]
            ]
            write_error_summary_sheet(wb, error_reasons, sheet_name="Errors")

        write_highlight_summary_sheet(
            wb,
            validated_df,
            highlight_cells.get("yellow"),
            sheet_name="Calculated Values",
        )
        if "Input Summary" in wb.sheetnames:
            del wb["Input Summary"]
        summary_ws = wb.create_sheet("Input Summary")
        summary_ws.append(["Property Name", owner_name])
        summary_ws.append(["Migration Date", migration_date])
        summary_ws.append(["Input Files", ", ".join(input_filenames)])
        header_font = Font(bold=False, size=14)
        value_font = Font(bold=True, size=14)
        for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row, max_col=2):
            for cell in row:
                cell.font = header_font if cell.column == 1 else value_font
        wb.save(tmp_path)

        if invalid_cells:
            # Print console report with Excel-style row numbers (1-based, excluding header).
            print("\nInvalid entries found:")
            for col, idx_list in invalid_cells.items():
                rows = [i + 2 for i in idx_list]
                print(f"- {col}: rows {rows}")

    filename = "final_output.xlsx"
    # FileResponse handles opening the file; we use os.remove in background cleanup.
    # BackgroundTask cleans up the temp file after the response is sent.
    background_task = BackgroundTask(
        lambda: (
            (os.path.exists(tmp_path) and os.remove(tmp_path)),
            (os.path.exists(mapping_tmp_path) and os.remove(mapping_tmp_path)),
        )
    )
    return FileResponse(
        path=tmp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        background=background_task,
    )
