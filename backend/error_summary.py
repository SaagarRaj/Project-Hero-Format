from __future__ import annotations

from typing import Dict, List

from openpyxl.styles import Border, Font, Side # type: ignore #typ


def _apply_table_style(ws, start_row: int, end_row: int, end_col: int) -> None:
    header_font = Font(bold=True)
    thin_side = Side(style="thin")
    full_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    for row in ws.iter_rows(min_row=start_row, max_row=end_row, max_col=end_col):
        for cell in row:
            if cell.row == start_row:
                cell.font = header_font
            cell.border = full_border


def write_error_summary_sheet(
    wb, invalid_reasons: List[Dict[str, object]], sheet_name: str = "Error Summary"
) -> None:
    if not invalid_reasons:
        return

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    ws.append(["column", "Space ID", "row_number", "reason", "value"])
    invalid_reasons_sorted = sorted(
        invalid_reasons, key=lambda r: (str(r.get("column", "")), r.get("row_index", 0))
    )
    for item in invalid_reasons_sorted:
        row_number = item.get("row_index", 0) + 2
        ws.append(
            [
                item.get("column", ""),
                item.get("space", ""),
                row_number,
                item.get("reason", ""),
                item.get("value", ""),
            ]
        )

    _apply_table_style(ws, 1, ws.max_row, 6)


def write_highlight_summary_sheet(
    wb,
    df,
    highlights: Dict[str, List[int]] | None,
    prev_values: Dict[str, Dict[int, object]] | None,
    sheet_name: str,
) -> None:
    if not highlights:
        return

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["column", "Space ID", "row_number", "value", "prev_value"])

    for col in sorted(highlights.keys()):
        idx_list = highlights.get(col) or []
        for idx in sorted(set(idx_list)):
            row_number = idx + 2
            value = df.at[idx, col] if col in df.columns else ""
            space_value = df.at[idx, "Space"] if "Space" in df.columns else ""
            prev_value = ""
            if prev_values and col in prev_values and idx in prev_values[col]:
                prev_value = prev_values[col][idx]
            ws.append([col, space_value, row_number, value, prev_value])

    _apply_table_style(ws, 1, ws.max_row, 5)
